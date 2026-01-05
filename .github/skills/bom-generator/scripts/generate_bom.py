#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl",
# ]
# ///
"""
BOM Generator - Creates Bill of Materials in multiple formats

Generates professional BOM spreadsheets (xlsx, csv) with:
- Component details and quantities
- Supplier links and pricing
- Category totals and grand total
- Formatted columns with formulas

Usage:
    uv run generate_bom.py --interactive
    uv run generate_bom.py --json input.json --output bom.xlsx
    uv run generate_bom.py --json input.json --format csv
"""

import argparse
import json
import csv
import sys
from dataclasses import dataclass, field
from typing import Optional, List
from pathlib import Path
from datetime import datetime

# Check for openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. xlsx output disabled.", file=sys.stderr)
    print("Install with: pip install openpyxl", file=sys.stderr)


# =============================================================================
# Component Database
# =============================================================================

COMPONENT_DATABASE = {
    # Microcontrollers
    "ESP32-DevKit": {
        "description": "ESP32 Development Board with WiFi/BLE",
        "category": "Microcontroller",
        "unit_price_usd": 8.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B08D5ZD528",
            "AliExpress": "https://aliexpress.com/item/32902307791.html"
        }
    },
    "Arduino-Uno-R3": {
        "description": "Arduino Uno R3 ATmega328P",
        "category": "Microcontroller",
        "unit_price_usd": 25.00,
        "package": "Module",
        "suppliers": {
            "Arduino.cc": "https://store.arduino.cc/products/arduino-uno-rev3",
            "Amazon": "https://amazon.com/dp/B008GRTSV6"
        }
    },
    "Arduino-Nano": {
        "description": "Arduino Nano V3 ATmega328P",
        "category": "Microcontroller",
        "unit_price_usd": 5.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B0713XK923"
        }
    },
    "Raspberry-Pi-Pico": {
        "description": "RP2040 Microcontroller Board",
        "category": "Microcontroller",
        "unit_price_usd": 4.00,
        "package": "Module",
        "suppliers": {
            "Raspberry Pi": "https://www.raspberrypi.com/products/raspberry-pi-pico/",
            "Adafruit": "https://www.adafruit.com/product/4864"
        }
    },
    
    # Sensors
    "BME280": {
        "description": "Temperature/Humidity/Pressure Sensor (I2C)",
        "category": "Sensor",
        "unit_price_usd": 3.50,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07KR24P6P",
            "AliExpress": "https://aliexpress.com/item/32849462236.html"
        }
    },
    "DHT22": {
        "description": "Temperature/Humidity Sensor",
        "category": "Sensor",
        "unit_price_usd": 4.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B073F472JL"
        }
    },
    "MPU6050": {
        "description": "6-Axis Accelerometer/Gyroscope (I2C)",
        "category": "Sensor",
        "unit_price_usd": 2.50,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B008BOPN40"
        }
    },
    "HC-SR04": {
        "description": "Ultrasonic Distance Sensor",
        "category": "Sensor",
        "unit_price_usd": 2.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07RGB4W8V"
        }
    },
    "PIR-HC-SR501": {
        "description": "PIR Motion Sensor",
        "category": "Sensor",
        "unit_price_usd": 1.50,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07KBWVJMP"
        }
    },
    
    # Displays
    "OLED-128x64": {
        "description": "0.96\" OLED Display 128x64 (I2C)",
        "category": "Display",
        "unit_price_usd": 5.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B08ZY4YBHL",
            "AliExpress": "https://aliexpress.com/item/32896971385.html"
        }
    },
    "LCD-16x2-I2C": {
        "description": "16x2 LCD Display with I2C Backpack",
        "category": "Display",
        "unit_price_usd": 6.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07S7PJYM6"
        }
    },
    "TFT-1.8-ST7735": {
        "description": "1.8\" TFT Display 128x160 (SPI)",
        "category": "Display",
        "unit_price_usd": 7.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B08FD643VZ"
        }
    },
    
    # Communication
    "NRF24L01": {
        "description": "2.4GHz Wireless Transceiver",
        "category": "Communication",
        "unit_price_usd": 2.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B00O9O868G"
        }
    },
    "LoRa-SX1276": {
        "description": "LoRa Radio Module 868/915MHz",
        "category": "Communication",
        "unit_price_usd": 12.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07V51M4VC"
        }
    },
    "GPS-NEO6M": {
        "description": "GPS Module with Antenna",
        "category": "Communication",
        "unit_price_usd": 10.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07P8YMVNT"
        }
    },
    
    # Power
    "TP4056": {
        "description": "LiPo Charging Module (USB-C)",
        "category": "Power",
        "unit_price_usd": 1.50,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07PZT3ZW2"
        }
    },
    "LM2596-DC-DC": {
        "description": "Buck Converter Adjustable DC-DC",
        "category": "Power",
        "unit_price_usd": 2.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B01GJ0SC2C"
        }
    },
    "18650-Battery": {
        "description": "18650 Li-ion Battery 3000mAh",
        "category": "Power",
        "unit_price_usd": 5.00,
        "package": "Cell",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07SXRXVF3"
        }
    },
    "LiPo-1000mAh": {
        "description": "LiPo Battery 3.7V 1000mAh",
        "category": "Power",
        "unit_price_usd": 6.00,
        "package": "Cell",
        "suppliers": {
            "Adafruit": "https://www.adafruit.com/product/258"
        }
    },
    
    # Actuators
    "Servo-SG90": {
        "description": "Micro Servo Motor SG90",
        "category": "Actuator",
        "unit_price_usd": 3.00,
        "package": "Unit",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07MLR1498"
        }
    },
    "Relay-5V": {
        "description": "5V Relay Module (1-Channel)",
        "category": "Actuator",
        "unit_price_usd": 2.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07BVXT1ZK"
        }
    },
    "Motor-Driver-L298N": {
        "description": "L298N Dual H-Bridge Motor Driver",
        "category": "Actuator",
        "unit_price_usd": 4.00,
        "package": "Module",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07BK1QL5T"
        }
    },
    
    # Passive Components
    "Resistor-Kit": {
        "description": "Resistor Assortment Kit (600pcs)",
        "category": "Passive",
        "unit_price_usd": 10.00,
        "package": "Kit",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B08FD1XVL6"
        }
    },
    "Capacitor-Kit": {
        "description": "Ceramic Capacitor Kit (300pcs)",
        "category": "Passive",
        "unit_price_usd": 12.00,
        "package": "Kit",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07PK78Y3H"
        }
    },
    "LED-Kit": {
        "description": "LED Assortment Kit (500pcs)",
        "category": "Passive",
        "unit_price_usd": 8.00,
        "package": "Kit",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07RBHSM6X"
        }
    },
    "Breadboard-830": {
        "description": "830 Point Solderless Breadboard",
        "category": "Passive",
        "unit_price_usd": 5.00,
        "package": "Unit",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B0135IQ0ZC"
        }
    },
    "Jumper-Wires-Kit": {
        "description": "Jumper Wire Kit M-M, M-F, F-F",
        "category": "Passive",
        "unit_price_usd": 7.00,
        "package": "Kit",
        "suppliers": {
            "Amazon": "https://amazon.com/dp/B07GD2BWPY"
        }
    }
}


@dataclass
class BOMItem:
    """Single item in the BOM"""
    reference: str  # e.g., "U1", "R1"
    name: str
    description: str
    quantity: int
    category: str
    unit_price_usd: float
    package: str = ""
    supplier: str = ""
    supplier_link: str = ""
    notes: str = ""
    
    @property
    def total_price(self) -> float:
        return self.quantity * self.unit_price_usd


@dataclass
class BOM:
    """Complete Bill of Materials"""
    project_name: str = "Untitled Project"
    version: str = "1.0"
    author: str = ""
    date: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))
    items: List[BOMItem] = field(default_factory=list)
    
    @property
    def total_cost(self) -> float:
        return sum(item.total_price for item in self.items)
    
    @property
    def item_count(self) -> int:
        return sum(item.quantity for item in self.items)
    
    def add_from_database(self, name: str, quantity: int = 1, 
                          reference: str = "", notes: str = "",
                          supplier: str = "") -> bool:
        """Add an item from the component database"""
        if name not in COMPONENT_DATABASE:
            return False
        
        comp = COMPONENT_DATABASE[name]
        
        # Select supplier
        suppliers = comp.get("suppliers", {})
        if supplier and supplier in suppliers:
            selected_supplier = supplier
            supplier_link = suppliers[supplier]
        elif suppliers:
            selected_supplier = list(suppliers.keys())[0]
            supplier_link = suppliers[selected_supplier]
        else:
            selected_supplier = ""
            supplier_link = ""
        
        item = BOMItem(
            reference=reference or f"X{len(self.items)+1}",
            name=name,
            description=comp["description"],
            quantity=quantity,
            category=comp["category"],
            unit_price_usd=comp["unit_price_usd"],
            package=comp.get("package", ""),
            supplier=selected_supplier,
            supplier_link=supplier_link,
            notes=notes
        )
        
        self.items.append(item)
        return True
    
    def add_custom(self, name: str, description: str, quantity: int,
                   category: str, unit_price: float, **kwargs) -> None:
        """Add a custom item not in the database"""
        item = BOMItem(
            reference=kwargs.get("reference", f"X{len(self.items)+1}"),
            name=name,
            description=description,
            quantity=quantity,
            category=category,
            unit_price_usd=unit_price,
            package=kwargs.get("package", ""),
            supplier=kwargs.get("supplier", ""),
            supplier_link=kwargs.get("supplier_link", ""),
            notes=kwargs.get("notes", "")
        )
        self.items.append(item)
    
    def to_csv(self, filename: str) -> None:
        """Export to CSV format"""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                "Reference", "Name", "Description", "Quantity", 
                "Category", "Unit Price ($)", "Total ($)", 
                "Package", "Supplier", "Supplier Link", "Notes"
            ])
            
            # Items
            for item in self.items:
                writer.writerow([
                    item.reference, item.name, item.description, item.quantity,
                    item.category, f"{item.unit_price_usd:.2f}", f"{item.total_price:.2f}",
                    item.package, item.supplier, item.supplier_link, item.notes
                ])
            
            # Total row
            writer.writerow([])
            writer.writerow(["", "", "", self.item_count, "", "", f"{self.total_cost:.2f}", "", "", "", ""])
    
    def to_xlsx(self, filename: str) -> bool:
        """Export to Excel xlsx format with formatting"""
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl not installed", file=sys.stderr)
            return False
        
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '$#,##0.00'
        
        # Project info
        ws['A1'] = "Project:"
        ws['B1'] = self.project_name
        ws['A2'] = "Version:"
        ws['B2'] = self.version
        ws['A3'] = "Author:"
        ws['B3'] = self.author
        ws['A4'] = "Date:"
        ws['B4'] = self.date
        
        ws['A1'].font = Font(bold=True)
        ws['A2'].font = Font(bold=True)
        ws['A3'].font = Font(bold=True)
        ws['A4'].font = Font(bold=True)
        
        # Header row
        headers = ["Ref", "Name", "Description", "Qty", "Category", 
                   "Unit Price", "Total", "Package", "Supplier", "Notes"]
        header_row = 6
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_num, item in enumerate(self.items, header_row + 1):
            data = [
                item.reference, item.name, item.description, item.quantity,
                item.category, item.unit_price_usd, item.total_price,
                item.package, item.supplier, item.notes
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                
                # Format currency columns
                if col in [6, 7]:
                    cell.number_format = currency_format
        
        # Total row
        total_row = header_row + len(self.items) + 2
        ws.cell(row=total_row, column=3, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=self.item_count).font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=7, value=self.total_cost)
        total_cell.font = Font(bold=True)
        total_cell.number_format = currency_format
        
        # Add formula for dynamic total
        formula_row = header_row + 1
        end_row = header_row + len(self.items)
        ws.cell(row=total_row, column=7, 
                value=f"=SUM(G{formula_row}:G{end_row})")
        
        # Adjust column widths
        column_widths = [8, 25, 40, 6, 15, 12, 12, 10, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Category summary sheet
        ws2 = wb.create_sheet(title="By Category")
        ws2['A1'] = "Category"
        ws2['B1'] = "Items"
        ws2['C1'] = "Total Cost"
        ws2['A1'].font = Font(bold=True)
        ws2['B1'].font = Font(bold=True)
        ws2['C1'].font = Font(bold=True)
        
        categories = {}
        for item in self.items:
            if item.category not in categories:
                categories[item.category] = {"count": 0, "cost": 0}
            categories[item.category]["count"] += item.quantity
            categories[item.category]["cost"] += item.total_price
        
        for row, (cat, data) in enumerate(sorted(categories.items()), 2):
            ws2.cell(row=row, column=1, value=cat)
            ws2.cell(row=row, column=2, value=data["count"])
            cost_cell = ws2.cell(row=row, column=3, value=data["cost"])
            cost_cell.number_format = currency_format
        
        # Save
        wb.save(filename)
        return True
    
    def to_markdown(self) -> str:
        """Generate markdown BOM table"""
        lines = [
            f"# Bill of Materials: {self.project_name}",
            "",
            f"**Version:** {self.version}",
            f"**Author:** {self.author}",
            f"**Date:** {self.date}",
            "",
            "## Components",
            "",
            "| Ref | Component | Description | Qty | Unit $ | Total $ | Supplier |",
            "|-----|-----------|-------------|-----|--------|---------|----------|"
        ]
        
        for item in self.items:
            supplier_cell = f"[{item.supplier}]({item.supplier_link})" if item.supplier_link else item.supplier
            lines.append(
                f"| {item.reference} | {item.name} | {item.description} | "
                f"{item.quantity} | ${item.unit_price_usd:.2f} | ${item.total_price:.2f} | {supplier_cell} |"
            )
        
        lines.extend([
            "",
            f"**Total Components:** {self.item_count}",
            f"**Estimated Cost:** ${self.total_cost:.2f}",
            "",
            "## By Category",
            ""
        ])
        
        categories = {}
        for item in self.items:
            if item.category not in categories:
                categories[item.category] = {"count": 0, "cost": 0}
            categories[item.category]["count"] += item.quantity
            categories[item.category]["cost"] += item.total_price
        
        for cat, data in sorted(categories.items()):
            lines.append(f"- **{cat}:** {data['count']} items, ${data['cost']:.2f}")
        
        return "\n".join(lines)
    
    def to_dict(self) -> dict:
        """Export to dictionary/JSON"""
        return {
            "project_name": self.project_name,
            "version": self.version,
            "author": self.author,
            "date": self.date,
            "items": [
                {
                    "reference": item.reference,
                    "name": item.name,
                    "description": item.description,
                    "quantity": item.quantity,
                    "category": item.category,
                    "unit_price_usd": item.unit_price_usd,
                    "total_price_usd": item.total_price,
                    "package": item.package,
                    "supplier": item.supplier,
                    "supplier_link": item.supplier_link,
                    "notes": item.notes
                }
                for item in self.items
            ],
            "summary": {
                "total_items": self.item_count,
                "total_cost_usd": round(self.total_cost, 2)
            }
        }
    
    @classmethod
    def from_json(cls, filepath: str) -> 'BOM':
        """Load BOM from JSON file"""
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        bom = cls(
            project_name=data.get("project_name", "Untitled"),
            version=data.get("version", "1.0"),
            author=data.get("author", ""),
            date=data.get("date", datetime.now().strftime("%Y-%m-%d"))
        )
        
        for item in data.get("items", []):
            if item.get("name") in COMPONENT_DATABASE:
                bom.add_from_database(
                    name=item["name"],
                    quantity=item.get("quantity", 1),
                    reference=item.get("reference", ""),
                    notes=item.get("notes", ""),
                    supplier=item.get("supplier", "")
                )
            else:
                bom.add_custom(
                    name=item.get("name", "Unknown"),
                    description=item.get("description", ""),
                    quantity=item.get("quantity", 1),
                    category=item.get("category", "Other"),
                    unit_price=item.get("unit_price_usd", 0),
                    reference=item.get("reference", ""),
                    supplier=item.get("supplier", ""),
                    supplier_link=item.get("supplier_link", ""),
                    notes=item.get("notes", "")
                )
        
        return bom


def interactive_mode():
    """Run BOM generator in interactive mode"""
    print("=" * 60)
    print("BOM Generator - Interactive Mode")
    print("=" * 60)
    print()
    
    project_name = input("Project name [My Project]: ").strip() or "My Project"
    version = input("Version [1.0]: ").strip() or "1.0"
    author = input("Author: ").strip()
    
    bom = BOM(project_name=project_name, version=version, author=author)
    
    print("\nAvailable components in database:")
    for i, name in enumerate(COMPONENT_DATABASE.keys(), 1):
        print(f"  {i:2}. {name}")
    
    print("\nEnter components (empty line to finish):")
    print("Format: <name or number> [quantity] [reference]")
    
    while True:
        line = input("\nComponent: ").strip()
        if not line:
            break
        
        parts = line.split()
        comp_input = parts[0]
        quantity = int(parts[1]) if len(parts) > 1 else 1
        reference = parts[2] if len(parts) > 2 else ""
        
        # Handle numeric input
        if comp_input.isdigit():
            idx = int(comp_input) - 1
            names = list(COMPONENT_DATABASE.keys())
            if 0 <= idx < len(names):
                comp_input = names[idx]
            else:
                print("Invalid number")
                continue
        
        if bom.add_from_database(comp_input, quantity, reference):
            comp = COMPONENT_DATABASE[comp_input]
            print(f"  Added: {quantity}x {comp_input} (${comp['unit_price_usd']:.2f} each)")
        else:
            # Custom component
            print(f"  '{comp_input}' not in database. Enter details:")
            desc = input("    Description: ").strip()
            category = input("    Category: ").strip() or "Other"
            price = float(input("    Unit price ($): ").strip() or "0")
            bom.add_custom(comp_input, desc, quantity, category, price, reference=reference)
            print(f"  Added custom: {quantity}x {comp_input}")
    
    # Show summary
    print("\n" + "=" * 60)
    print(bom.to_markdown())
    
    # Export
    print("\nExport options:")
    print("1. Excel (.xlsx)")
    print("2. CSV (.csv)")
    print("3. Markdown (.md)")
    print("4. JSON (.json)")
    print("5. All formats")
    
    choice = input("\nExport format [1]: ").strip() or "1"
    base_name = project_name.replace(" ", "_")
    
    if choice in ["1", "5"]:
        if bom.to_xlsx(f"{base_name}_BOM.xlsx"):
            print(f"Saved: {base_name}_BOM.xlsx")
    
    if choice in ["2", "5"]:
        bom.to_csv(f"{base_name}_BOM.csv")
        print(f"Saved: {base_name}_BOM.csv")
    
    if choice in ["3", "5"]:
        with open(f"{base_name}_BOM.md", 'w') as f:
            f.write(bom.to_markdown())
        print(f"Saved: {base_name}_BOM.md")
    
    if choice in ["4", "5"]:
        with open(f"{base_name}_BOM.json", 'w') as f:
            json.dump(bom.to_dict(), f, indent=2)
        print(f"Saved: {base_name}_BOM.json")


def main():
    parser = argparse.ArgumentParser(description="BOM Generator for Embedded Projects")
    parser.add_argument("--interactive", "-i", action="store_true", help="Interactive mode")
    parser.add_argument("--json", "-j", type=str, help="Load from JSON file")
    parser.add_argument("--output", "-o", type=str, help="Output filename")
    parser.add_argument("--format", "-f", type=str, choices=["xlsx", "csv", "md", "json"],
                        default="xlsx", help="Output format")
    parser.add_argument("--list", "-l", action="store_true", help="List component database")
    
    args = parser.parse_args()
    
    if args.list:
        print("Component Database:")
        print("-" * 80)
        for name, data in COMPONENT_DATABASE.items():
            print(f"{name:25} ${data['unit_price_usd']:6.2f}  {data['description']}")
        return
    
    if args.interactive:
        interactive_mode()
        return
    
    if args.json:
        bom = BOM.from_json(args.json)
        output = args.output or f"{bom.project_name.replace(' ', '_')}_BOM.{args.format}"
        
        if args.format == "xlsx":
            bom.to_xlsx(output)
        elif args.format == "csv":
            bom.to_csv(output)
        elif args.format == "md":
            with open(output, 'w') as f:
                f.write(bom.to_markdown())
        elif args.format == "json":
            with open(output, 'w') as f:
                json.dump(bom.to_dict(), f, indent=2)
        
        print(f"Generated: {output}")
        return
    
    parser.print_help()


if __name__ == "__main__":
    main()
